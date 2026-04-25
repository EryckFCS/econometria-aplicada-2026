from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from pathlib import Path

# Configuración
PROJECT_ROOT = Path(
    "/home/erick-fcs/Documentos/universidad/07_Ciclo/septimo_ciclo/econometria-aplicada-2026"
)
excel_path = (
    PROJECT_ROOT / "data" / "raw" / "excel" / "MASTER_FINAL_CONSOLIDADO_API_LOCAL.xlsx"
)

# Mapeo de Colores por Responsable
RESPONSABLES_COLORS = {
    "Erick Condoy": "DCE6F1",  # Azul
    "Marvin Valdivieso": "EBF1DE",  # Verde
    "Abel Espinosa": "FDE9D9",  # Naranja
    "Helen Beltran": "F2DCDB",  # Rosa/Rojo
    "Erik Moreno": "E4DFEC",  # Púrpura
    "WB": "F2F2F2",  # Gris (API)
    "World Bank": "F2F2F2",
}

# Mapeo de Variables a Responsables (para el diccionario)
VAR_TO_RESP = {
    "gei_total_exclulucf_MtCO2e": "Erick Condoy",
    "renew_cons_pct_tfec": "Erick Condoy",
    "wgi_va_est": "Erick Condoy",
    "wgi_ge_est": "Erick Condoy",
    "informalidad": "Marvin Valdivieso",
    "tecnologia": "Marvin Valdivieso",
    "calidad_institucional": "Marvin Valdivieso",
    "gobernanza": "Marvin Valdivieso",
    "indice_riesgo_climatico": "Abel Espinosa",
    "PIBpc": "Abel Espinosa",
    "renta_rec_nat": "Abel Espinosa",
    "impuestos_ambientales": "Abel Espinosa",
    "protec_der_lab": "Abel Espinosa",
    "climate_altering_land_index": "Abel Espinosa",
    "poblacion": "Helen Beltran",
    "incertidumbre_energetica": "Helen Beltran",
    "superficie_forestal": "Erik Moreno",
    "vab_manufactura": "Erik Moreno",
    "vab_agricultura": "Erik Moreno",
    "vab_servicios": "Erik Moreno",
    "huella_ecologica": "Erik Moreno",
    "impuesto_medioambiental_moreno": "Erik Moreno",
    "co2_emissions_moreno": "Erik Moreno",
}


def apply_colors():
    print(f"🎨 Aplicando colores al diccionario en: {excel_path.name}")

    wb = load_workbook(excel_path)
    if "Diccionario" not in wb.sheetnames:
        print("❌ No se encontró la hoja 'Diccionario'")
        return

    ws = wb["Diccionario"]

    # Estilo de cabecera
    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Aplicar colores por fila basándose en la variable
    # Asumimos que la columna 'variable_local' es la primera (A)
    for row in range(2, ws.max_row + 1):
        var_name = ws.cell(row=row, column=1).value
        resp = VAR_TO_RESP.get(var_name, "WB")
        color = RESPONSABLES_COLORS.get(resp, "FFFFFF")

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill

    # Ajustar anchos
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 40

    wb.save(excel_path)
    print("✅ Colores aplicados exitosamente.")


if __name__ == "__main__":
    apply_colors()
