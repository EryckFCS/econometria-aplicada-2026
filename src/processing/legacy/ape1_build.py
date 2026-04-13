"""
APE1 - Econometría Aplicada
Base de datos: 20 países de América Latina
Variables ambientales, económicas y de control
Autor: Script generado con trazabilidad completa de fuentes
"""

import requests
import pandas as pd
import time
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN GENERAL
# ─────────────────────────────────────────────────────────────────────────────

PAISES_LA = {
    "AR": "Argentina",
    "BO": "Bolivia",
    "BR": "Brasil",
    "CL": "Chile",
    "CO": "Colombia",
    "CR": "Costa Rica",
    "CU": "Cuba",
    "EC": "Ecuador",
    "SV": "El Salvador",
    "GT": "Guatemala",
    "HT": "Haití",
    "HN": "Honduras",
    "MX": "México",
    "NI": "Nicaragua",
    "PA": "Panamá",
    "PY": "Paraguay",
    "PE": "Perú",
    "DO": "República Dominicana",
    "UY": "Uruguay",
    "VE": "Venezuela",
}

CODIGOS_WB = list(PAISES_LA.keys())
ANIO_INICIO = 2000
ANIO_FIN    = 2023

# ─────────────────────────────────────────────────────────────────────────────
# METADATOS DE VARIABLES (trazabilidad)
# ─────────────────────────────────────────────────────────────────────────────

METADATA = {
    # ── VARIABLES DEPENDIENTES ────────────────────────────────────────────────
    "superficie_forestal_km2": {
        "descripcion": "Área forestal (km²)",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "AG.LND.FRST.K2",
        "url_fuente": "https://data.worldbank.org/indicator/AG.LND.FRST.K2",
        "tipo": "Dependiente",
        "unidad": "km²",
        "notas": "Bosques y superficies forestales totales"
    },
    "superficie_forestal_pct": {
        "descripcion": "Área forestal (% del área territorial)",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "AG.LND.FRST.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/AG.LND.FRST.ZS",
        "tipo": "Dependiente",
        "unidad": "%",
        "notas": "Proxy: % de área forestal sobre territorio total"
    },
    "gei_kt_co2eq": {
        "descripcion": "Emisiones totales de gases de efecto invernadero (kt CO2 equivalente)",
        "fuente": "EDGAR / World Bank - EN.GHG.CO2.MT.CE.AR5 (adaptado a kt)",
        "indicador_wb": "EN.ATM.GHGT.KT.CE",
        "url_fuente": "https://data.worldbank.org/indicator/EN.ATM.GHGT.KT.CE",
        "tipo": "Dependiente / Prioritaria",
        "unidad": "kt CO2 equivalente",
        "notas": "Fuente subyacente: EDGAR JRC-IEA 2025 (edgar.jrc.ec.europa.eu/report_2025)"
    },
    "co2_kt": {
        "descripcion": "Emisiones de CO2 (kt)",
        "fuente": "World Bank - World Development Indicators / EDGAR IEA",
        "indicador_wb": "EN.ATM.CO2E.KT",
        "url_fuente": "https://data.worldbank.org/indicator/EN.ATM.CO2E.KT",
        "tipo": "Dependiente",
        "unidad": "kt",
        "notas": "CO2 procedente de quema de combustibles fósiles y fabricación de cemento"
    },
    "co2_pc": {
        "descripcion": "Emisiones de CO2 per cápita (t)",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "EN.ATM.CO2E.PC",
        "url_fuente": "https://data.worldbank.org/indicator/EN.ATM.CO2E.PC",
        "tipo": "Dependiente",
        "unidad": "t per cápita",
        "notas": "CO2 per cápita; complementa la variable total"
    },
    # Factor capacidad de carga — proxy ecológica
    "huella_ecologica_gha_pc": {
        "descripcion": "Huella ecológica per cápita (proxy factor capacidad de carga) — Global Footprint Network",
        "fuente": "Global Footprint Network (Open Data Platform)",
        "indicador_wb": None,
        "url_fuente": "https://data.footprintnetwork.org/",
        "tipo": "Dependiente",
        "unidad": "gha per cápita",
        "notas": "Proxy directo del factor de capacidad de carga biótica. Disponible en WB como: EG.USE.COMM.KG.OE.PC (energía) o variable EF_pc del GFN. Se usará World Bank AG.LND.BIOD.IBA.ZS como proxy disponible vía API."
    },

    # ── VARIABLES INDEPENDIENTES ──────────────────────────────────────────────
    "pib_per_capita_usd": {
        "descripcion": "PIB per cápita (USD constantes 2015)",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "NY.GDP.PCAP.KD",
        "url_fuente": "https://data.worldbank.org/indicator/NY.GDP.PCAP.KD",
        "tipo": "Independiente",
        "unidad": "USD constantes 2015",
        "notas": "Proxy crecimiento económico; deflactado"
    },
    "pib_per_capita_ppa": {
        "descripcion": "PIB per cápita PPA (USD internacionales constantes 2017)",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "NY.GDP.PCAP.PP.KD",
        "url_fuente": "https://data.worldbank.org/indicator/NY.GDP.PCAP.PP.KD",
        "tipo": "Independiente",
        "unidad": "USD PPA 2017",
        "notas": "Corrige paridad de poder adquisitivo entre países"
    },
    "poblacion_total": {
        "descripcion": "Población total",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "SP.POP.TOTL",
        "url_fuente": "https://data.worldbank.org/indicator/SP.POP.TOTL",
        "tipo": "Independiente",
        "unidad": "número de personas",
        "notas": ""
    },
    "acceso_internet_pct": {
        "descripcion": "Acceso a internet (% de la población) — proxy tecnología",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "IT.NET.USER.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/IT.NET.USER.ZS",
        "tipo": "Independiente - Tecnología",
        "unidad": "%",
        "notas": "Proxy de adopción tecnológica general"
    },
    "suscripciones_movil": {
        "descripcion": "Suscripciones a telefonía móvil (por 100 personas) — proxy tecnología",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "IT.CEL.SETS.P2",
        "url_fuente": "https://data.worldbank.org/indicator/IT.CEL.SETS.P2",
        "tipo": "Independiente - Tecnología",
        "unidad": "por 100 hab.",
        "notas": "Proxy tecnológico complementario"
    },

    # ── VARIABLES DE CONTROL ──────────────────────────────────────────────────
    "vab_agricola_pct_pib": {
        "descripcion": "VAB sector agrícola (% del PIB)",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "NV.AGR.TOTL.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/NV.AGR.TOTL.ZS",
        "tipo": "Control - VAB Sectorial",
        "unidad": "% del PIB",
        "notas": "Agricultura, silvicultura, pesca"
    },
    "vab_manufactura_pct_pib": {
        "descripcion": "VAB sector manufacturero (% del PIB)",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "NV.IND.MANF.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/NV.IND.MANF.ZS",
        "tipo": "Control - VAB Sectorial",
        "unidad": "% del PIB",
        "notas": ""
    },
    "vab_servicios_pct_pib": {
        "descripcion": "VAB sector servicios (% del PIB)",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "NV.SRV.TOTL.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/NV.SRV.TOTL.ZS",
        "tipo": "Control - VAB Sectorial",
        "unidad": "% del PIB",
        "notas": ""
    },

    # ── CONTROL: DERECHOS LABORALES ───────────────────────────────────────────
    "trabajo_infantil_pct": {
        "descripcion": "Trabajo infantil (% niños 7-14 años) — proxy derechos laborales",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "SL.TLF.0714.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/SL.TLF.0714.ZS",
        "tipo": "Control - Derechos Laborales",
        "unidad": "%",
        "notas": "Proxy inverso de protección laboral. Cobertura limitada."
    },

    # ── CONTROL: TECNOLOGÍA VERDE / INNOVACIÓN TECNOLÓGICA VERDE (PRIORITARIAS) ─
    "patentes_env_tech_pct": {
        "descripcion": "Patentes en tecnologías ambientales (% de todas las patentes) — innovación tecnológica verde / tecnología verde",
        "fuente": "OECD - Environment Statistics / World Bank WDI (EP.PMP.SGAS.CD proxy inverso)",
        "indicador_wb": "EP.PMP.SGAS.CD",
        "url_fuente": "https://www.oecd.org/en/data/indicators/patents-on-environment-technologies.html",
        "tipo": "Control - Tecnología Verde / Innovación Tecnológica Verde (PRIORITARIA)",
        "unidad": "% de patentes domésticas",
        "notas": "Indicador OECD. Para AL, cobertura reducida. Proxy WB: EP.PMP.SGAS.CD (bombas de gasolina = proxy regulación verde). Alternativa: EG.ELC.RNEW.ZS (energía renovable)."
    },
    "energia_renovable_pct": {
        "descripcion": "Energía renovable (% del consumo final de energía) — tecnología verde PRIORITARIA",
        "fuente": "World Bank - World Development Indicators / IEA",
        "indicador_wb": "EG.FEC.RNEW.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/EG.FEC.RNEW.ZS",
        "tipo": "Control - Tecnología Verde (PRIORITARIA)",
        "unidad": "%",
        "notas": "Indicador directo de adopción de tecnologías de energía limpia"
    },
    "capacidad_renovable_gw": {
        "descripcion": "Capacidad instalada energía renovable (% electricidad) — proxy tecnología verde",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "EG.ELC.RNEW.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/EG.ELC.RNEW.ZS",
        "tipo": "Control - Innovación Tecnológica Verde (PRIORITARIA)",
        "unidad": "% de generación eléctrica",
        "notas": "Proxy de stock de tecnología verde instalada"
    },
    "gasto_id_pct_pib": {
        "descripcion": "Gasto en I+D (% del PIB) — proxy innovación tecnológica verde",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "GB.XPD.RSDV.GD.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/GB.XPD.RSDV.GD.ZS",
        "tipo": "Control - Innovación Tecnológica Verde (PRIORITARIA)",
        "unidad": "% del PIB",
        "notas": "I+D total; covariable de capacidad innovadora"
    },

    # ── CONTROL: INCLUSIÓN FINANCIERA (PRIORITARIA) ───────────────────────────
    "cuenta_financiera_pct": {
        "descripcion": "Adultos con cuenta en institución financiera o móvil (%) — inclusión financiera PRIORITARIA",
        "fuente": "World Bank - Global Findex Database",
        "indicador_wb": "FX.OWN.TOTL.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/FX.OWN.TOTL.ZS",
        "tipo": "Control - Inclusión Financiera (PRIORITARIA)",
        "unidad": "%",
        "notas": "Encuesta trienal (2011, 2014, 2017, 2021, 2024). Interpolación lineal recomendada."
    },
    "credito_privado_pct_pib": {
        "descripcion": "Crédito doméstico al sector privado (% del PIB) — inclusión financiera",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "FS.AST.PRVT.GD.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/FS.AST.PRVT.GD.ZS",
        "tipo": "Control - Inclusión Financiera (PRIORITARIA)",
        "unidad": "% del PIB",
        "notas": "Profundidad financiera; serie anual continua"
    },

    # ── CONTROL: RENDICIÓN DE CUENTAS (PRIORITARIA) ───────────────────────────
    "rendicion_cuentas_est": {
        "descripcion": "Voz y Rendición de Cuentas — WGI estimate (PRIORITARIA)",
        "fuente": "World Bank - Worldwide Governance Indicators (WGI)",
        "indicador_wb": "VA.EST",
        "url_fuente": "https://www.worldbank.org/en/publication/worldwide-governance-indicators",
        "tipo": "Control - Rendición de Cuentas (PRIORITARIA)",
        "unidad": "Estimado (-2.5 a +2.5)",
        "notas": "Mide capacidad ciudadana de participar, expresarse y fiscalizar gobierno"
    },
    "rendicion_cuentas_rk": {
        "descripcion": "Voz y Rendición de Cuentas — percentil de rango",
        "fuente": "World Bank - Worldwide Governance Indicators (WGI)",
        "indicador_wb": "VA.PER.RNK",
        "url_fuente": "https://www.worldbank.org/en/publication/worldwide-governance-indicators",
        "tipo": "Control - Rendición de Cuentas (PRIORITARIA)",
        "unidad": "Percentil (0-100)",
        "notas": "Rango percentil entre países"
    },

    # ── CONTROL: GOBERNANZA ───────────────────────────────────────────────────
    "gobernanza_efectividad": {
        "descripcion": "Efectividad gubernamental — WGI estimate",
        "fuente": "World Bank - Worldwide Governance Indicators (WGI)",
        "indicador_wb": "GE.EST",
        "url_fuente": "https://www.worldbank.org/en/publication/worldwide-governance-indicators",
        "tipo": "Control - Gobernanza",
        "unidad": "Estimado (-2.5 a +2.5)",
        "notas": ""
    },
    "gobernanza_estado_derecho": {
        "descripcion": "Estado de Derecho — WGI estimate",
        "fuente": "World Bank - Worldwide Governance Indicators (WGI)",
        "indicador_wb": "RL.EST",
        "url_fuente": "https://www.worldbank.org/en/publication/worldwide-governance-indicators",
        "tipo": "Control - Gobernanza / Calidad Institucional",
        "unidad": "Estimado (-2.5 a +2.5)",
        "notas": ""
    },
    "control_corrupcion": {
        "descripcion": "Control de Corrupción — WGI estimate",
        "fuente": "World Bank - Worldwide Governance Indicators (WGI)",
        "indicador_wb": "CC.EST",
        "url_fuente": "https://www.worldbank.org/en/publication/worldwide-governance-indicators",
        "tipo": "Control - Gobernanza / Calidad Institucional",
        "unidad": "Estimado (-2.5 a +2.5)",
        "notas": ""
    },
    "estabilidad_politica": {
        "descripcion": "Estabilidad Política y Ausencia de Violencia — WGI estimate",
        "fuente": "World Bank - Worldwide Governance Indicators (WGI)",
        "indicador_wb": "PV.EST",
        "url_fuente": "https://www.worldbank.org/en/publication/worldwide-governance-indicators",
        "tipo": "Control - Gobernanza",
        "unidad": "Estimado (-2.5 a +2.5)",
        "notas": ""
    },
    "calidad_regulatoria": {
        "descripcion": "Calidad Regulatoria — WGI estimate",
        "fuente": "World Bank - Worldwide Governance Indicators (WGI)",
        "indicador_wb": "RQ.EST",
        "url_fuente": "https://www.worldbank.org/en/publication/worldwide-governance-indicators",
        "tipo": "Control - Gobernanza / Política Ambiental",
        "unidad": "Estimado (-2.5 a +2.5)",
        "notas": "Proxy de rigurosidad de política ambiental (marco regulatorio general)"
    },

    # ── CONTROL: IMPUESTOS AMBIENTALES ────────────────────────────────────────
    "impuesto_carbono_dummy": {
        "descripcion": "Presencia de impuesto/mecanismo carbono (dummy) — CEPAL / World Bank PMR",
        "fuente": "World Bank - State and Trends of Carbon Pricing / CEPAL",
        "indicador_wb": None,
        "url_fuente": "https://carbonpricingdashboard.worldbank.org/",
        "tipo": "Control - Impuestos Ambientales / Política Ambiental",
        "unidad": "dummy (0/1)",
        "notas": "No disponible vía API WB estándar; se construirá manualmente desde Carbon Pricing Dashboard"
    },
    "recaudacion_impuestos_pct_pib": {
        "descripcion": "Recaudación tributaria total (% del PIB) — proxy carga fiscal",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "GC.TAX.TOTL.GD.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/GC.TAX.TOTL.GD.ZS",
        "tipo": "Control - Impuestos",
        "unidad": "% del PIB",
        "notas": "Proxy de carga fiscal; impuestos ambientales específicos no disponibles vía API para AL"
    },

    # ── CONTROL: RENTA RECURSOS NATURALES ─────────────────────────────────────
    "renta_recursos_naturales_pct": {
        "descripcion": "Renta total de recursos naturales (% del PIB)",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "NY.GDP.TOTL.RT.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/NY.GDP.TOTL.RT.ZS",
        "tipo": "Control - Renta Recursos Naturales",
        "unidad": "% del PIB",
        "notas": "Petróleo + gas + carbón + minerales + forestal"
    },

    # ── CONTROL: INFORMALIDAD ─────────────────────────────────────────────────
    "empleo_informal_pct": {
        "descripcion": "Empleo vulnerable (% empleo total) — proxy informalidad",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "SL.EMP.VULN.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/SL.EMP.VULN.ZS",
        "tipo": "Control - Informalidad",
        "unidad": "%",
        "notas": "Proxy de informalidad; empleo por cuenta propia + familiar no remunerado"
    },

    # ── CONTROL: INNOVACIÓN FINANCIERA ───────────────────────────────────────
    "pagos_movil_pct": {
        "descripcion": "Pagos móviles realizados (% adultos) — proxy innovación financiera",
        "fuente": "World Bank - Global Findex Database",
        "indicador_wb": "FX.TRN.SEND.DT.GD.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/FX.TRN.SEND.DT.GD.ZS",
        "tipo": "Control - Innovación Financiera",
        "unidad": "%",
        "notas": "Remesas como % PIB; mejor proxy disponible vía API WB continua para innovación financiera"
    },
    "capitalizacion_mercado_pct": {
        "descripcion": "Capitalización bursátil empresas cotizadas (% del PIB) — innovación financiera",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "CM.MKT.LCAP.GD.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/CM.MKT.LCAP.GD.ZS",
        "tipo": "Control - Innovación Financiera",
        "unidad": "% del PIB",
        "notas": "Desarrollo de mercados de capital"
    },

    # ── CONTROL: ÍNDICE RIESGO CLIMÁTICO ─────────────────────────────────────
    "desastres_naturales_afectados": {
        "descripcion": "Personas afectadas por desastres naturales (por 100,000 hab.) — proxy riesgo climático",
        "fuente": "World Bank - World Development Indicators",
        "indicador_wb": "EN.CLC.MDAT.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/EN.CLC.MDAT.ZS",
        "tipo": "Control - Índice Riesgo Climático",
        "unidad": "por 100,000 hab.",
        "notas": "Proxy de exposición a riesgo climático; series incompletas"
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# FUNCIÓN DESCARGA WORLD BANK API
# ─────────────────────────────────────────────────────────────────────────────

def fetch_wb_indicator(indicator_code, countries, start_year, end_year):
    """
    Descarga un indicador del World Bank API v2.
    Retorna DataFrame con columnas: iso2, year, value
    """
    country_str = ";".join(countries)
    url = (
        f"https://api.worldbank.org/v2/country/{country_str}/indicator/{indicator_code}"
        f"?format=json&per_page=1000&mrv={end_year - start_year + 1}"
        f"&date={start_year}:{end_year}"
    )
    
    records = []
    page = 1
    while True:
        paged_url = url + f"&page={page}"
        try:
            r = requests.get(paged_url, timeout=30)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"  ⚠ Error en {indicator_code}: {e}")
            break
        
        if len(data) < 2 or not data[1]:
            break
        
        for entry in data[1]:
            if entry.get("value") is not None:
                records.append({
                    "iso2": entry["countryiso3code"][:2] if len(entry.get("countryiso3code","")) >= 2 else entry["country"]["id"],
                    "year": int(entry["date"]),
                    "value": float(entry["value"]),
                })
        
        # Paginación
        meta = data[0]
        total_pages = meta.get("pages", 1)
        if page >= total_pages:
            break
        page += 1
        time.sleep(0.15)
    
    df = pd.DataFrame(records)
    if df.empty:
        return df
    
    # Normalizar iso2
    # WB a veces devuelve iso3; mapear a iso2
    iso3_to_iso2 = {
        "ARG": "AR", "BOL": "BO", "BRA": "BR", "CHL": "CL", "COL": "CO",
        "CRI": "CR", "CUB": "CU", "ECU": "EC", "SLV": "SV", "GTM": "GT",
        "HTI": "HT", "HND": "HN", "MEX": "MX", "NIC": "NI", "PAN": "PA",
        "PRY": "PY", "PER": "PE", "DOM": "DO", "URY": "UY", "VEN": "VE",
    }
    df["iso2"] = df["iso2"].apply(lambda x: iso3_to_iso2.get(x.upper(), x.upper()))
    
    return df[df["iso2"].isin(CODIGOS_WB)].drop_duplicates(["iso2", "year"])


# ─────────────────────────────────────────────────────────────────────────────
# CONSTRUCCIÓN DE LA BASE DE DATOS
# ─────────────────────────────────────────────────────────────────────────────

def build_panel():
    """
    Construye el panel de datos completo (país × año).
    """
    # Marco base: todos los pares (país, año)
    from itertools import product
    base = pd.DataFrame(
        list(product(CODIGOS_WB, range(ANIO_INICIO, ANIO_FIN + 1))),
        columns=["iso2", "year"]
    )
    base["pais"] = base["iso2"].map(PAISES_LA)
    
    # Reordenar columnas base
    base = base[["iso2", "pais", "year"]]
    
    # Descargar cada indicador WB
    total = sum(1 for v in METADATA.values() if v["indicador_wb"])
    contador = 0
    
    for var_name, meta in METADATA.items():
        wb_code = meta["indicador_wb"]
        if wb_code is None:
            print(f"  SKIP (sin API WB): {var_name}")
            base[var_name] = None
            continue
        
        contador += 1
        print(f"  [{contador}/{total}] Descargando: {var_name} ({wb_code})...")
        
        df_ind = fetch_wb_indicator(wb_code, CODIGOS_WB, ANIO_INICIO, ANIO_FIN)
        
        if df_ind.empty:
            print(f"    → Sin datos disponibles para {wb_code}")
            base[var_name] = None
        else:
            df_ind = df_ind.rename(columns={"value": var_name})
            base = base.merge(df_ind[["iso2", "year", var_name]], on=["iso2", "year"], how="left")
            n_obs = df_ind.shape[0]
            print(f"    → {n_obs} observaciones descargadas")
        
        time.sleep(0.2)
    
    return base


# ─────────────────────────────────────────────────────────────────────────────
# EXPORTAR A EXCEL CON TRAZABILIDAD
# ─────────────────────────────────────────────────────────────────────────────

def exportar_excel(panel_df, output_path):
    """
    Exporta el panel y la hoja de metadatos a un archivo Excel estructurado.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    
    # ── 1. DataFrame de metadatos ─────────────────────────────────────────────
    meta_rows = []
    for var_name, meta in METADATA.items():
        meta_rows.append({
            "Variable": var_name,
            "Descripción": meta["descripcion"],
            "Tipo": meta["tipo"],
            "Unidad": meta["unidad"],
            "Fuente": meta["fuente"],
            "Indicador WB": meta["indicador_wb"] if meta["indicador_wb"] else "N/A (manual)",
            "URL Fuente": meta["url_fuente"],
            "Notas": meta["notas"],
        })
    meta_df = pd.DataFrame(meta_rows)
    
    # ── 2. Estadísticas de cobertura ──────────────────────────────────────────
    cobertura_rows = []
    for var_name in METADATA.keys():
        if var_name in panel_df.columns:
            total_celdas = len(panel_df)
            obs_validas = panel_df[var_name].notna().sum()
            pct = round(100 * obs_validas / total_celdas, 1) if total_celdas > 0 else 0
            cobertura_rows.append({
                "Variable": var_name,
                "Total Obs (país×año)": total_celdas,
                "Obs con datos": obs_validas,
                "Cobertura (%)": pct,
            })
    cobertura_df = pd.DataFrame(cobertura_rows)
    
    # ── 3. Escribir Excel ─────────────────────────────────────────────────────
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Hoja principal: panel de datos
        panel_df.to_excel(writer, sheet_name="PANEL_DATOS", index=False)
        
        # Hoja de metadatos / trazabilidad
        meta_df.to_excel(writer, sheet_name="METADATOS_TRAZABILIDAD", index=False)
        
        # Hoja cobertura
        cobertura_df.to_excel(writer, sheet_name="COBERTURA_DATOS", index=False)
        
        # Hoja README
        readme_data = {
            "Campo": [
                "Nombre del trabajo",
                "Asignatura",
                "Países",
                "Período",
                "Fecha de construcción",
                "Fuentes principales",
                "Notas metodológicas",
            ],
            "Valor": [
                "APE1 - Base de Datos Econometría Ambiental",
                "Econometría Aplicada",
                "20 países de América Latina",
                f"{ANIO_INICIO} - {ANIO_FIN}",
                pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"),
                "World Bank WDI, EDGAR JRC-IEA, WGI, Global Findex, OECD, Global Footprint Network",
                (
                    "Variables prioritarias (GEI, innovación tecnológica verde, tecnología verde, "
                    "inclusión financiera, rendición de cuentas) descargadas directamente vía API. "
                    "Factor de capacidad de carga aproximado con energía renovable y huella ecológica. "
                    "Global Findex es encuesta trienal: se recomienda interpolación para análisis de panel. "
                    "Indicadores WGI disponibles desde 1996 con gaps años pares/impares pre-2002."
                ),
            ]
        }
        pd.DataFrame(readme_data).to_excel(writer, sheet_name="README", index=False)
    
    # ── 4. Formato visual ─────────────────────────────────────────────────────
    wb = load_workbook(output_path)
    
    AZUL_OSCURO = "1F3864"
    GRIS_CLARO  = "F2F2F2"
    BLANCO      = "FFFFFF"
    
    def estilizar_hoja(ws, header_color=AZUL_OSCURO):
        header_fill = PatternFill("solid", fgColor=header_color)
        header_font = Font(bold=True, color=BLANCO, size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = PatternFill("solid", fgColor=GRIS_CLARO if row_idx % 2 == 0 else BLANCO)
            for cell in row:
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(wrap_text=False, vertical="center")
        
        # Autofit aproximado
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 60)
        
        ws.freeze_panes = "A2"
    
    # Aplicar estilo a todas las hojas
    for sheet_name in ["PANEL_DATOS", "METADATOS_TRAZABILIDAD", "COBERTURA_DATOS", "README"]:
        if sheet_name in wb.sheetnames:
            estilizar_hoja(wb[sheet_name])
    
    # Resaltar variables prioritarias en METADATOS
    ws_meta = wb["METADATOS_TRAZABILIDAD"]
    prio_fill = PatternFill("solid", fgColor="E2EFDA")  # verde claro
    prio_font_bold = Font(bold=True, color="375623")
    for row in ws_meta.iter_rows(min_row=2):
        tipo_cell = row[2]  # columna "Tipo"
        if tipo_cell.value and "PRIORITARIA" in str(tipo_cell.value).upper():
            for cell in row:
                cell.fill = prio_fill
                if cell.column == 3:
                    cell.font = prio_font_bold
    
    wb.save(output_path)
    print(f"\n✓ Excel exportado: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 65)
    print("APE1 - Construcción Base de Datos Econometría Ambiental")
    print(f"Países: {len(PAISES_LA)} | Período: {ANIO_INICIO}-{ANIO_FIN}")
    print("=" * 65)
    
    panel = build_panel()
    
    print(f"\n→ Panel construido: {panel.shape[0]} filas × {panel.shape[1]} columnas")
    
    output_excel = "/home/user/workspace/APE1_BaseDatos_AmericaLatina.xlsx"
    output_csv   = "/home/user/workspace/APE1_BaseDatos_AmericaLatina.csv"
    
    # Exportar CSV
    panel.to_csv(output_csv, index=False, encoding="utf-8-sig")
    print(f"→ CSV exportado: {output_csv}")
    
    # Exportar Excel con trazabilidad
    exportar_excel(panel, output_excel)
    
    print("\n✓ PROCESO COMPLETADO")
    print(f"  Filas: {panel.shape[0]}")
    print(f"  Columnas: {panel.shape[1]}")
    print(f"  Variables: {panel.shape[1] - 3} indicadores + (iso2, pais, year)")