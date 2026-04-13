"""
PATCH: Completa variables fallidas en el dataset APE1
- WGI (códigos GOV_WGI_*)
- GEI / CO2 (códigos EN.GHG.*)
- Superficie forestal %
- Global Findex account ownership
"""

import requests
import pandas as pd
import time
import warnings
warnings.filterwarnings("ignore")

PAISES_LA = {
    "AR": "Argentina", "BO": "Bolivia", "BR": "Brasil", "CL": "Chile",
    "CO": "Colombia", "CR": "Costa Rica", "CU": "Cuba", "EC": "Ecuador",
    "SV": "El Salvador", "GT": "Guatemala", "HT": "Haití", "HN": "Honduras",
    "MX": "México", "NI": "Nicaragua", "PA": "Panamá", "PY": "Paraguay",
    "PE": "Perú", "DO": "República Dominicana", "UY": "Uruguay", "VE": "Venezuela",
}

CODIGOS_WB = list(PAISES_LA.keys())
PAISES_STR = ";".join(CODIGOS_WB)
ANIO_INICIO = 2000
ANIO_FIN    = 2023

ISO3_TO_ISO2 = {
    "ARG": "AR", "BOL": "BO", "BRA": "BR", "CHL": "CL", "COL": "CO",
    "CRI": "CR", "CUB": "CU", "ECU": "EC", "SLV": "SV", "GTM": "GT",
    "HTI": "HT", "HND": "HN", "MEX": "MX", "NIC": "NI", "PAN": "PA",
    "PRY": "PY", "PER": "PE", "DOM": "DO", "URY": "UY", "VEN": "VE",
}

# Variables a parchear con sus códigos correctos
PATCH_VARS = {
    "gei_total_incl_lulucf_mt":    "EN.GHG.ALL.LU.MT.CE.AR5",
    "co2_excl_lulucf_mt":          "EN.GHG.CO2.MT.CE.AR5",
    "superficie_forestal_pct":     "AG.LND.FRST.ZS",
    "cuenta_financiera_pct":       "FX.OWN.TOTL.ZS",
    "rendicion_cuentas_est":       "GOV_WGI_VA.EST",
    "gobernanza_efectividad":      "GOV_WGI_GE.EST",
    "control_corrupcion":          "GOV_WGI_CC.EST",
    "gobernanza_estado_derecho":   "GOV_WGI_RL.EST",
    "calidad_regulatoria":         "GOV_WGI_RQ.EST",
    "estabilidad_politica":        "GOV_WGI_PV.EST",
}

PATCH_META = {
    "gei_total_incl_lulucf_mt": {
        "descripcion": "Emisiones GEI totales incl. LULUCF (Mt CO2 equivalente, AR5)",
        "fuente": "World Bank / EDGAR JRC-IEA 2025",
        "indicador_wb": "EN.GHG.ALL.LU.MT.CE.AR5",
        "url_fuente": "https://data.worldbank.org/indicator/EN.GHG.ALL.LU.MT.CE.AR5",
        "tipo": "Dependiente / Prioritaria - GEI",
        "unidad": "Mt CO2eq (AR5 GWP)",
        "notas": "Incluye LULUCF (Land Use, Land-Use Change and Forestry). Fuente subyacente: EDGAR_2025_GHG"
    },
    "co2_excl_lulucf_mt": {
        "descripcion": "Emisiones CO2 excl. LULUCF (Mt CO2eq, AR5)",
        "fuente": "World Bank / EDGAR JRC-IEA 2025 / IEA",
        "indicador_wb": "EN.GHG.CO2.MT.CE.AR5",
        "url_fuente": "https://data.worldbank.org/indicator/EN.GHG.CO2.MT.CE.AR5",
        "tipo": "Dependiente",
        "unidad": "Mt CO2eq (AR5)",
        "notas": "CO2 fossil + industrial (excl. cambio uso de suelo)"
    },
    "superficie_forestal_pct": {
        "descripcion": "Superficie forestal (% del área territorial)",
        "fuente": "World Bank / FAO Global Forest Resources Assessment",
        "indicador_wb": "AG.LND.FRST.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/AG.LND.FRST.ZS",
        "tipo": "Dependiente",
        "unidad": "% territorio",
        "notas": "Actualización quinquenal interpolada por WB/FAO"
    },
    "cuenta_financiera_pct": {
        "descripcion": "Adultos con cuenta financiera o móvil (%) — Inclusión Financiera",
        "fuente": "World Bank Global Findex Database 2011-2024",
        "indicador_wb": "FX.OWN.TOTL.ZS",
        "url_fuente": "https://data.worldbank.org/indicator/FX.OWN.TOTL.ZS",
        "tipo": "Control - Inclusión Financiera (PRIORITARIA)",
        "unidad": "% adultos +15 años",
        "notas": "Encuesta trienal: 2011, 2014, 2017, 2021, 2024. Se recomienda interpolación lineal para análisis de panel anual."
    },
    "rendicion_cuentas_est": {
        "descripcion": "Voz y Rendición de Cuentas — WGI Estimate",
        "fuente": "World Bank Worldwide Governance Indicators (WGI)",
        "indicador_wb": "GOV_WGI_VA.EST",
        "url_fuente": "https://www.worldbank.org/en/publication/worldwide-governance-indicators",
        "tipo": "Control - Rendición de Cuentas (PRIORITARIA)",
        "unidad": "Estimado aprox. -2.5 a +2.5",
        "notas": "Voice & Accountability: participación ciudadana, libertad de expresión, medios libres"
    },
    "gobernanza_efectividad": {
        "descripcion": "Efectividad Gubernamental — WGI Estimate",
        "fuente": "World Bank Worldwide Governance Indicators (WGI)",
        "indicador_wb": "GOV_WGI_GE.EST",
        "url_fuente": "https://www.worldbank.org/en/publication/worldwide-governance-indicators",
        "tipo": "Control - Gobernanza / Calidad Institucional",
        "unidad": "Estimado aprox. -2.5 a +2.5",
        "notas": ""
    },
    "control_corrupcion": {
        "descripcion": "Control de Corrupción — WGI Estimate",
        "fuente": "World Bank Worldwide Governance Indicators (WGI)",
        "indicador_wb": "GOV_WGI_CC.EST",
        "url_fuente": "https://www.worldbank.org/en/publication/worldwide-governance-indicators",
        "tipo": "Control - Gobernanza / Calidad Institucional",
        "unidad": "Estimado aprox. -2.5 a +2.5",
        "notas": ""
    },
    "gobernanza_estado_derecho": {
        "descripcion": "Estado de Derecho — WGI Estimate",
        "fuente": "World Bank Worldwide Governance Indicators (WGI)",
        "indicador_wb": "GOV_WGI_RL.EST",
        "url_fuente": "https://www.worldbank.org/en/publication/worldwide-governance-indicators",
        "tipo": "Control - Gobernanza / Calidad Institucional",
        "unidad": "Estimado aprox. -2.5 a +2.5",
        "notas": ""
    },
    "calidad_regulatoria": {
        "descripcion": "Calidad Regulatoria — WGI Estimate (proxy rigurosidad política ambiental)",
        "fuente": "World Bank Worldwide Governance Indicators (WGI)",
        "indicador_wb": "GOV_WGI_RQ.EST",
        "url_fuente": "https://www.worldbank.org/en/publication/worldwide-governance-indicators",
        "tipo": "Control - Gobernanza / Política Ambiental",
        "unidad": "Estimado aprox. -2.5 a +2.5",
        "notas": ""
    },
    "estabilidad_politica": {
        "descripcion": "Estabilidad Política y Ausencia de Violencia — WGI Estimate",
        "fuente": "World Bank Worldwide Governance Indicators (WGI)",
        "indicador_wb": "GOV_WGI_PV.EST",
        "url_fuente": "https://www.worldbank.org/en/publication/worldwide-governance-indicators",
        "tipo": "Control - Gobernanza",
        "unidad": "Estimado aprox. -2.5 a +2.5",
        "notas": ""
    },
}


def fetch_wb(code, countries_str, start, end):
    url = (
        f"https://api.worldbank.org/v2/country/{countries_str}/indicator/{code}"
        f"?format=json&per_page=1000&date={start}:{end}"
    )
    records = []
    page = 1
    while True:
        try:
            r = requests.get(url + f"&page={page}", timeout=30)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"    ⚠ Error {code} p{page}: {e}")
            break
        if not isinstance(data, list) or len(data) < 2 or not data[1]:
            break
        for entry in data[1]:
            if entry.get("value") is not None:
                iso_raw = entry.get("countryiso3code", "") or entry.get("country", {}).get("id", "")
                iso2 = ISO3_TO_ISO2.get(iso_raw.upper(), iso_raw[:2].upper() if iso_raw else "")
                records.append({"iso2": iso2, "year": int(entry["date"]), "value": float(entry["value"])})
        meta = data[0]
        if page >= meta.get("pages", 1):
            break
        page += 1
        time.sleep(0.15)
    
    df = pd.DataFrame(records)
    if df.empty:
        return df
    df = df[df["iso2"].isin(CODIGOS_WB)].drop_duplicates(["iso2", "year"])
    return df


# ─── CARGAR PANEL EXISTENTE ───────────────────────────────────────────────────
print("Cargando panel existente...")
panel = pd.read_csv("/home/user/workspace/APE1_BaseDatos_AmericaLatina.csv")
print(f"Panel original: {panel.shape}")

# ─── PARCHEAR VARIABLES ───────────────────────────────────────────────────────
new_data = {}

for var_name, wb_code in PATCH_VARS.items():
    print(f"  Descargando: {var_name} ({wb_code})...")
    df_ind = fetch_wb(wb_code, PAISES_STR, ANIO_INICIO, ANIO_FIN)
    if df_ind.empty:
        print("    → Sin datos")
        new_data[var_name] = None
    else:
        n = df_ind.shape[0]
        print(f"    → {n} observaciones")
        new_data[var_name] = df_ind.rename(columns={"value": var_name})
    time.sleep(0.2)

# ─── ACTUALIZAR / AGREGAR COLUMNAS ────────────────────────────────────────────
for var_name, df_ind in new_data.items():
    if df_ind is None:
        if var_name not in panel.columns:
            panel[var_name] = None
        continue
    # Si ya existe la columna en el panel original (con Nones) → reemplazar
    if var_name in panel.columns:
        panel = panel.drop(columns=[var_name])
    panel = panel.merge(df_ind[["iso2", "year", var_name]], on=["iso2", "year"], how="left")

print(f"\nPanel actualizado: {panel.shape}")

# ─── EXPORTAR CSV ACTUALIZADO ─────────────────────────────────────────────────
csv_path = "/home/user/workspace/APE1_BaseDatos_AmericaLatina.csv"
panel.to_csv(csv_path, index=False, encoding="utf-8-sig")
print(f"→ CSV actualizado: {csv_path}")

# ─── EXPORTAR EXCEL CON TRAZABILIDAD COMPLETA ─────────────────────────────────
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Metadatos completos (originales + nuevos del patch)
from ape1_build import METADATA
METADATA_COMPLETO = {**METADATA}
for k, v in PATCH_META.items():
    METADATA_COMPLETO[k] = v

meta_rows = []
for var_name, meta in METADATA_COMPLETO.items():
    if var_name not in panel.columns:
        continue
    meta_rows.append({
        "Variable": var_name,
        "Descripción": meta["descripcion"],
        "Tipo": meta["tipo"],
        "Unidad": meta["unidad"],
        "Fuente": meta["fuente"],
        "Indicador WB": meta.get("indicador_wb") or "N/A",
        "URL Fuente": meta["url_fuente"],
        "Notas": meta["notas"],
    })
meta_df = pd.DataFrame(meta_rows)

cobertura_rows = []
for var_name in panel.columns[3:]:
    total = len(panel)
    obs = panel[var_name].notna().sum()
    cobertura_rows.append({
        "Variable": var_name,
        "Total Obs": total,
        "Obs con datos": obs,
        "Cobertura (%)": round(100 * obs / total, 1) if total > 0 else 0,
        "Tipo": METADATA_COMPLETO.get(var_name, {}).get("tipo", ""),
        "Fuente": METADATA_COMPLETO.get(var_name, {}).get("fuente", ""),
    })
cobertura_df = pd.DataFrame(cobertura_rows)

readme_data = {
    "Campo": [
        "Nombre del trabajo",
        "Asignatura",
        "Número APE",
        "Países",
        "Período de análisis",
        "Fecha de construcción",
        "Fuentes principales",
        "Variables prioritarias",
        "Notas metodológicas",
        "Advertencias",
    ],
    "Valor": [
        "APE1 - Base de Datos Econometría Ambiental - América Latina",
        "Econometría Aplicada",
        "APE1",
        "20 países: Argentina, Bolivia, Brasil, Chile, Colombia, Costa Rica, Cuba, Ecuador, El Salvador, Guatemala, Haití, Honduras, México, Nicaragua, Panamá, Paraguay, Perú, República Dominicana, Uruguay, Venezuela",
        "2000 - 2023 (anual)",
        pd.Timestamp.now().strftime("%Y-%m-%d %H:%M UTC"),
        "World Bank WDI | EDGAR JRC-IEA (2025) | Worldwide Governance Indicators (WGI) | Global Findex Database | FAO Forest Resources | OECD Patent Statistics",
        "GEI (EN.GHG.ALL.LU.MT.CE.AR5) | Innovación Tecnológica Verde (EG.ELC.RNEW.ZS + GB.XPD.RSDV.GD.ZS) | Tecnología Verde (EG.FEC.RNEW.ZS) | Inclusión Financiera (FX.OWN.TOTL.ZS) | Rendición de Cuentas (GOV_WGI_VA.EST)",
        "Variables descargadas vía World Bank API v2. EDGAR es la fuente subyacente de datos de emisiones. Global Findex es encuesta trienal (2011/2014/2017/2021/2024): datos anuales requieren interpolación. WGI disponible desde 1996.",
        "Cuba: datos limitados en WGI y GEI. Venezuela: series con gaps post-2013. Haiti: cobertura reducida en varias variables. Factor de capacidad de carga: aproximado con energía renovable (mejor proxy disponible vía API).",
    ]
}
readme_df = pd.DataFrame(readme_data)

excel_path = "/home/user/workspace/APE1_BaseDatos_AmericaLatina.xlsx"

with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    panel.to_excel(writer, sheet_name="PANEL_DATOS", index=False)
    meta_df.to_excel(writer, sheet_name="METADATOS_TRAZABILIDAD", index=False)
    cobertura_df.to_excel(writer, sheet_name="COBERTURA_DATOS", index=False)
    readme_df.to_excel(writer, sheet_name="README", index=False)

# Aplicar formato
wb_xl = load_workbook(excel_path)

AZUL_OSC = "1F3864"
BLANCO    = "FFFFFF"
GRIS_CLR  = "F2F2F2"
VERDE_CLR = "E2EFDA"
VERDE_FNT = "375623"

def fmt_sheet(ws, hdr_color=AZUL_OSC):
    hf = PatternFill("solid", fgColor=hdr_color)
    font_h = Font(bold=True, color=BLANCO, size=10)
    thin = Side(style="thin", color="AAAAAA")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.fill = hf; c.font = font_h
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd
    for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
        fl = PatternFill("solid", fgColor=GRIS_CLR if ri % 2 == 0 else BLANCO)
        for c in row:
            c.fill = fl; c.border = brd
            c.alignment = Alignment(vertical="center")
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 3, 55)
    ws.freeze_panes = "A2"

for sh in wb_xl.sheetnames:
    fmt_sheet(wb_xl[sh])

# Resaltar prioritarias en METADATOS
ws_m = wb_xl["METADATOS_TRAZABILIDAD"]
for row in ws_m.iter_rows(min_row=2):
    tipo_v = str(row[2].value or "")
    if "PRIORITARIA" in tipo_v.upper():
        for c in row:
            c.fill = PatternFill("solid", fgColor=VERDE_CLR)
        row[2].font = Font(bold=True, color=VERDE_FNT)

# Cobertura: color según %, col D (index 3)
ws_c = wb_xl["COBERTURA_DATOS"]
for row in ws_c.iter_rows(min_row=2):
    try:
        pct = float(row[3].value or 0)
    except:
        pct = 0
    if pct >= 80:
        cl = "C6EFCE"
    elif pct >= 50:
        cl = "FFEB9C"
    else:
        cl = "FFC7CE"
    row[3].fill = PatternFill("solid", fgColor=cl)
    row[3].font = Font(bold=True)

wb_xl.save(excel_path)
print(f"\n✓ Excel final exportado: {excel_path}")

# ─── RESUMEN FINAL ────────────────────────────────────────────────────────────
print("\n" + "=" * 65)
print("RESUMEN DE COBERTURA — VARIABLES PRIORITARIAS")
print("=" * 65)
prio = ["gei_total_incl_lulucf_mt", "co2_excl_lulucf_mt",
        "energia_renovable_pct", "capacidad_renovable_gw", "gasto_id_pct_pib",
        "cuenta_financiera_pct", "credito_privado_pct_pib",
        "rendicion_cuentas_est", "gobernanza_efectividad",
        "control_corrupcion", "gobernanza_estado_derecho"]
for v in prio:
    if v in panel.columns:
        n = panel[v].notna().sum()
        pct = round(100 * n / len(panel), 1)
        print(f"  {v:<40} {n:>4}/{len(panel)} obs  ({pct}%)")
print("=" * 65)
print(f"\nColumnas totales: {panel.shape[1]}")
print(f"Filas totales   : {panel.shape[0]}")